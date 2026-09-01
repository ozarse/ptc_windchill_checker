"""Excel-export comparison check — Windchill "where used" products vs a published list.

Validates the products that use Config PDPs (the ``used_by`` targets stored by
``sync relationships``) against a bulk export from the IFU publishing website:

  1. Every Windchill product must appear in the export (by product number).
  2. For products present in both, the export's IFU numbers must equal the IFU
     PDP numbers reachable in Windchill (product -> its Config PDPs -> uses).
  3. Export products that no Config PDP is used by are reported as orphans.

Products do not need to be synced into the ``objects`` table — the ``used_by``
relationship rows carry their real Number. The Excel file is read at check time;
if it is absent the check emits a single skip row instead of failing.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from ptc_syncer_ingestion.db import get_object_by_id, get_objects_by_type, get_relationship_targets
from ptc_syncer_ingestion.models import CheckResult, ExcelCompareCheck

log = logging.getLogger(__name__)

# Cap how many numbers are listed inline in a message.
_MAX_LISTED = 8


def _fmt(numbers: set[str]) -> str:
    listed = sorted(numbers)
    if len(listed) > _MAX_LISTED:
        return ", ".join(listed[:_MAX_LISTED]) + f" (+{len(listed) - _MAX_LISTED} more)"
    return ", ".join(listed)


def _cell_str(value) -> str:
    """Normalize a cell value to a stripped string ('' for empty cells)."""
    if value is None:
        return ""
    return str(value).strip()


def load_export(
    path: Path, sheet: str | None, product_column: str, ifu_column: str, separator: str
) -> dict[str, set[str]]:
    """Read the export into ``{product_number: {ifu_number, ...}}``.

    Headers are taken from the first row. A product appearing on several rows
    gets the union of its IFU sets. Raises ValueError when the sheet or the
    configured columns cannot be found.
    """
    from openpyxl import load_workbook  # deferred: keep CLI startup fast

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in {path} (sheets: {wb.sheetnames})")
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]

        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            raise ValueError(f"Export {path} has no header row")
        header_index = {_cell_str(h): i for i, h in enumerate(headers) if _cell_str(h)}
        missing = [c for c in (product_column, ifu_column) if c not in header_index]
        if missing:
            raise ValueError(
                f"Column(s) {missing} not found in {path}; headers are {sorted(header_index)}"
            )
        p_idx = header_index[product_column]
        i_idx = header_index[ifu_column]

        products: dict[str, set[str]] = {}
        for row in rows:
            product = _cell_str(row[p_idx]) if p_idx < len(row) else ""
            if not product:
                continue
            raw = _cell_str(row[i_idx]) if i_idx < len(row) else ""
            ifus = {part.strip() for part in raw.split(separator) if part.strip()}
            products.setdefault(product, set()).update(ifus)
        return products
    finally:
        wb.close()


def _windchill_products(conn, config_type: str, ifu_type: str) -> dict[str, dict]:
    """Collect products (used_by targets of ``config_type`` records) from the DB.

    Returns ``{product_number: {"id", "configs", "ifus"}}`` where ``ifus`` is the
    union of ``ifu_type`` numbers used by that product's config parents.
    """
    products: dict[str, dict] = {}
    for config in get_objects_by_type(conn, config_type):
        config_label = config.number or config.id
        ifus: set[str] = set()
        for target_id, _ in get_relationship_targets(conn, config.id, "uses"):
            obj = get_object_by_id(conn, target_id)
            if obj is not None and obj.type_name == ifu_type and obj.number:
                ifus.add(obj.number)
        for product_id, product_number in get_relationship_targets(conn, config.id, "used_by"):
            entry = products.setdefault(
                product_number or product_id, {"id": product_id, "configs": set(), "ifus": set()}
            )
            entry["configs"].add(config_label)
            entry["ifus"].update(ifus)
    return products


def run_excel_compare_check(conn, check: ExcelCompareCheck) -> list[CheckResult]:
    """Run the three-way comparison. See the module docstring."""
    now = datetime.now(timezone.utc).isoformat()

    def _row(source_id, target_id, source_attr, target_attr, source_value, target_value,
             passed, message, status=""):
        return CheckResult(
            check_name=check.name,
            source_object_id=source_id,
            target_object_id=target_id,
            source_attr=source_attr,
            target_attr=target_attr,
            source_value=source_value,
            target_value=target_value,
            passed=passed,
            status=status,
            message=message,
            checked_at=now,
        )

    path = Path(check.file)
    if not path.exists():
        return [_row(
            "", "", "", "", None, None, True,
            f"SKIP: export file not found: {path} — place the published-IFU export there "
            f"(or update 'file' in checks.json) and re-run",
            status="skip",
        )]

    windchill = _windchill_products(conn, check.type, check.ifu_type)
    if not windchill:
        return [_row(
            "", "", "", "", None, None, True,
            f"SKIP: no 'used_by' relationships stored for any {check.type} — "
            f"run 'ptc_syncer sync relationships' first",
            status="skip",
        )]

    export = load_export(path, check.sheet, check.product_column, check.ifu_column, check.ifu_separator)
    results: list[CheckResult] = []

    for number in sorted(windchill):
        info = windchill[number]
        configs = _fmt(info["configs"])
        if number not in export:
            results.append(_row(
                info["id"], "MISSING", "Number", check.product_column, number, None, False,
                f"FAIL: product {number} (used by {check.type} {configs}) "
                f"not found in published export",
            ))
            continue

        wc_ifus: set[str] = info["ifus"]
        ex_ifus = export[number]
        unpublished = wc_ifus - ex_ifus
        unknown = ex_ifus - wc_ifus
        passed = not unpublished and not unknown
        if passed:
            message = f"PASS: product {number} published with matching IFUs ({_fmt(ex_ifus) or 'none'})"
        else:
            details = []
            if unpublished:
                details.append(f"in Windchill but not published: {_fmt(unpublished)}")
            if unknown:
                details.append(f"published but not in Windchill: {_fmt(unknown)}")
            message = f"FAIL: product {number} IFU mismatch — " + "; ".join(details)
        results.append(_row(
            info["id"], number, f"{check.ifu_type} numbers", check.ifu_column,
            check.ifu_separator.join(sorted(wc_ifus)),
            check.ifu_separator.join(sorted(ex_ifus)),
            passed, message,
        ))

    for number in sorted(set(export) - set(windchill)):
        results.append(_row(
            "EXPORT", number, "", check.product_column, None, number, False,
            f"FAIL: published product {number} is not used by any {check.type} in Windchill",
        ))

    return results
